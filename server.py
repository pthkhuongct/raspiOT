from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from threading import Thread
from time import sleep
from datetime import datetime
import snap7
from snap7.exceptions import Snap7Exception
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
app.secret_key = 'your_secret_key'
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

voltage_db3 = 0.0
current_db3 = 0.0
kwh_db3 = 0.0
voltage_db4 = 0.0
current_db4 = 0.0
kwh_db4 = 0.0

class User(UserMixin):
    def __init__(self, id):
        self.id = id

@login_manager.user_loader
def load_user(user_id):
    return User(user_id)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if username == 'admin' and password == 'admin':
            user = User(id='admin')
            login_user(user)
            return redirect(url_for('index'))
        else:
            flash('Sai tên đăng nhập hoặc mật khẩu')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    return render_template('index.html', voltage_db3=voltage_db3, current_db3=current_db3, kwh_db3=kwh_db3, voltage_db4=voltage_db4, current_db4=current_db4, kwh_db4=kwh_db4)

@app.route('/data')
@login_required
def data():
    return jsonify(voltage_db3=voltage_db3, current_db3=current_db3, kwh_db3=kwh_db3, voltage_db4=voltage_db4, current_db4=current_db4, kwh_db4=kwh_db4)

def read_plc_values():
    global voltage_db3, current_db3, kwh_db3, voltage_db4, current_db4, kwh_db4

    plc = snap7.client.Client()

    try:
        plc.connect("192.168.1.199", 0, 1)
        if plc.get_connected():
            print("Kết nối thành công đến PLC.")
        else:
            print("Kết nối đến PLC thất bại.")
            return
    except Snap7Exception as e:
        print(f"Lỗi khi kết nối đến PLC: {e}")
        return

    while True:
        try:
            seft_db3 = plc.db_read(3, 0, 12)  # đọc 12 byte từ DB3
            voltage_db3 = snap7.util.get_real(seft_db3, 0)
            current_db3 = snap7.util.get_real(seft_db3, 4)
            kwh_db3 = snap7.util.get_real(seft_db3, 8)

            seft_db4 = plc.db_read(4, 0, 12)  # đọc 12 byte từ DB4
            voltage_db4 = snap7.util.get_real(seft_db4, 0)
            current_db4 = snap7.util.get_real(seft_db4, 4)
            kwh_db4 = snap7.util.get_real(seft_db4, 8)

            print(f"Đọc DB3: Điện áp {voltage_db3} V, Dòng điện {current_db3} A, kWh {kwh_db3}, Đọc DB4: Điện áp {voltage_db4} V, Dòng điện {current_db4} A, kWh {kwh_db4}")
        except Snap7Exception as e:
            print(f"Lỗi khi đọc dữ liệu từ PLC: {e}")
        except Exception as e:
            print(f"Một lỗi không mong muốn đã xảy ra: {e}")
        
        sleep(1)

def update_excel():
    file_path = 'database.xlsx'
    
    if not os.path.exists(file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Dữ liệu điện áp và dòng điện"
        sheet.append(["Thời gian", "Điện áp đồng hồ 1 (V)", "Dòng điện đồng hồ 1 (A)", "kWh đồng hồ 1", "Điện áp đồng hồ 2 (V)", "Dòng điện đồng hồ 2 (A)", "kWh đồng hồ 2"])
        workbook.save(file_path)

    while True:
        try:
            if os.path.exists(file_path):
                workbook = load_workbook(file_path)
                sheet = workbook.active
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                sheet.append([timestamp, voltage_db3, current_db3, kwh_db3, voltage_db4, current_db4, kwh_db4])
                workbook.save(file_path)
                print(f"Cập nhật {file_path} vào lúc {timestamp}")
        except Exception as e:
            print(f"Lỗi khi cập nhật tệp Excel: {e}")
        
        sleep(60)

if __name__ == '__main__':
    plc_thread = Thread(target=read_plc_values)
    plc_thread.daemon = True
    plc_thread.start()

    excel_thread = Thread(target=update_excel)
    excel_thread.daemon = True
    excel_thread.start()

    app.run(host='0.0.0.0', port=5000)
